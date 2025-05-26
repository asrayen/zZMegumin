import requests
import json
import time
import random
import datetime
import os
from openpyxl import Workbook
from openpyxl import load_workbook

cookies = {
    'sajssdk_2015_cross_new_user': '1',
    'sensorsdata2015jssdkcross':
    '%7B%22distinct_id%22%3A%2218c9b56bc57b98-0d64390783d3e48-4c657b58-921600-18c9b56bc581af3%22%2C%22first_id%22%3A%22%22%2C%22props%22%3A%7B%22%24latest_traffic_source_type%22%3A%22%E7%9B%B4%E6%8E%A5%E6%B5%81%E9%87%8F%22%2C%22%24latest_search_keyword%22%3A%22%E6%9C%AA%E5%8F%96%E5%88%B0%E5%80%BC_%E7%9B%B4%E6%8E%A5%E6%89%93%E5%BC%80%22%2C%22%24latest_referrer%22%3A%22%22%7D%2C%22identities%22%3A%22eyIkaWRlbnRpdHlfY29va2llX2lkIjoiMThjOWI1NmJjNTdiOTgtMGQ2NDM5MDc4M2QzZTQ4LTRjNjU3YjU4LTkyMTYwMC0xOGM5YjU2YmM1ODFhZjMifQ%3D%3D%22%2C%22history_login_id%22%3A%7B%22name%22%3A%22%22%2C%22value%22%3A%22%22%7D%2C%22%24device_id%22%3A%2218c9b56bc57b98-0d64390783d3e48-4c657b58-921600-18c9b56bc581af3%22%7D',
    'Hm_lvt_07aa95427da600fc217b1133c1e84e5b': '1703413202',
    'Hm_lpvt_07aa95427da600fc217b1133c1e84e5b': '1703413294',
}
headers = {
    'Accept': 'application/json, text/plain, */*',
    'Accept-Language': 'zh-CN,zh;q=0.9',
    'Connection': 'keep-alive',
    'Content-Type': 'application/json',
    # Requests sorts cookies= alphabetically
    # 'Cookie': 'sajssdk_2015_cross_new_user=1; sensorsdata2015jssdkcross=%7B%22distinct_id%22%3A%2218c9b56bc57b98-0d64390783d3e48-4c657b58-921600-18c9b56bc581af3%22%2C%22first_id%22%3A%22%22%2C%22props%22%3A%7B%22%24latest_traffic_source_type%22%3A%22%E7%9B%B4%E6%8E%A5%E6%B5%81%E9%87%8F%22%2C%22%24latest_search_keyword%22%3A%22%E6%9C%AA%E5%8F%96%E5%88%B0%E5%80%BC_%E7%9B%B4%E6%8E%A5%E6%89%93%E5%BC%80%22%2C%22%24latest_referrer%22%3A%22%22%7D%2C%22identities%22%3A%22eyIkaWRlbnRpdHlfY29va2llX2lkIjoiMThjOWI1NmJjNTdiOTgtMGQ2NDM5MDc4M2QzZTQ4LTRjNjU3YjU4LTkyMTYwMC0xOGM5YjU2YmM1ODFhZjMifQ%3D%3D%22%2C%22history_login_id%22%3A%7B%22name%22%3A%22%22%2C%22value%22%3A%22%22%7D%2C%22%24device_id%22%3A%2218c9b56bc57b98-0d64390783d3e48-4c657b58-921600-18c9b56bc581af3%22%7D; Hm_lvt_07aa95427da600fc217b1133c1e84e5b=1703413202; Hm_lpvt_07aa95427da600fc217b1133c1e84e5b=1703413294',
    'Origin': 'http://movie.mtime.com',
    'Referer': 'http://movie.mtime.com/',
    'User-Agent':
    'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36 Edg/120.0.0.0',
    'X-Mtime-Wap-CheckValue': 'mtime',
}
filepath = '星际穿越(时光).xlsx'


def getinfo(i):
    allinfo = []
    try:
        params = (
            ('tt', '1703413293645'),
            ('movieId', '51119'),
            ('pageIndex', str(i)),
            ('pageSize', '20'),
            ('orderType', '1'),
        )
        r = requests.get(
            'http://front-gateway.mtime.com/library/movie/comment.api',
            params=params,
            cookies=cookies,
            headers=headers,
            verify=False)
        r.raise_for_status()
        r.encoding = 'utf-8'
        datas = json.loads(r.text)
        for item in datas['data']['list']:
            id = item['userId']
            nickName = item['nickname']
            content = item['content']
            score = item['rating']
            if score == '10':
                score = '5'
            elif score[0] == '9' and len(score) > 1:
                score = '5'
            elif score == '9':
                score = '4.5'
            elif score[0] == '8' and len(score) > 1:
                score = '4.5'
            elif score == '8':
                score = '4'
            elif score[0] == '7' and len(score) > 1:
                score = '4'
            elif score == '7':
                score = '3.5'
            elif score[0] == '6' and len(score) > 1:
                score = '3.5'
            elif score == '6':
                score = '3'
            elif score[0] == '5' and len(score) > 1:
                score = '3'
            elif score == '5':
                score = '2.5'
            elif score[0] == '4' and len(score) > 1:
                score = '2.5'
            elif score == '4':
                score = '2'
            elif score[0] == '3' and len(score) > 1:
                score = '2'
            elif score == '3':
                score = '1.5'
            elif score[0] == '2' and len(score) > 1:
                score = '1.5'
            elif score == '2':
                score = '1'
            elif score[0] == '1' and len(score) > 1:
                score = '1'
            elif score == '1':
                score = '0.5'
            elif score[0] == '0' and len(score) > 1:
                score = '0.5'
            else:
                score = '0'
            startTime = item['commentTime']
            comments = [id, nickName, content, score, startTime]
            print(comments)
            allinfo.append(comments)
        return allinfo
    except:
        if (len(allinfo) > 0):
            return allinfo


def insertexcel(filepath, allinfo):
    try:
        if not os.path.exists(filepath):
            tableTitle = ['id', '用户名', '评论', '评分', '评论时间']
            wb = Workbook()
            ws = wb.active
            ws.title = 'sheet1'
            ws.append(tableTitle)
            wb.save(filepath)
            time.sleep(0.5)
        wb = load_workbook(filepath)
        ws = wb.active
        ws.title = 'sheet1'
        for info in allinfo:
            ws.append(info)
        wb.save(filepath)
        print('文件已更新')
    except:
        print('文件更新失败')


for i in range(1, 600):
    time.sleep(random.uniform(0.1, 0.3))
    allinfo = getinfo(i)
    insertexcel(filepath, allinfo)
