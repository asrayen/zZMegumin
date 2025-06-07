import requests
import json
import time
import random
import datetime
import os
from openpyxl import Workbook
from openpyxl import load_workbook

start_time = '2023-12-24 00:00:00'
end_time = '2014-11-24 00:00:00'
filmid = 1298384
headers = {
    'accept':
    'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8',
    'user-agent':
    'Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/68.0.3440.106 Safari/537.36',
}

filepath = 'test.xlsx'


def getinfo(url, headers):
    allinfo = []
    try:
        r = requests.get(url, headers=headers)
        r.raise_for_status()
        r.encoding = 'utf-8'
        datas = json.loads(r.text)
        for items in datas['cmts']:
            id = items['id'],
            nickName = items['nickName'],
            cityName = items['cityName'] if 'cityName' in items else '',
            content = items['content'].replace('\n', ' '),
            score = items['score'],
            startTime = items['startTime'],
            gender = items['gender'] if 'gender' in items else ''
            comments = [
                id[0], nickName[0], cityName[0], content[0], score[0],
                gender[0], startTime
            ]
            print(comments)
            allinfo.append(comments)
        return allinfo
    except:
        if (len(allinfo) > 0):
            return allinfo


def insertexcel(filepath, allinfo):
    try:
        if not os.path.exists(filepath):
            tableTitle = ['id', '用户名', '城市', '评论', '评分', '性别', '评论时间']
            wb = Workbook()
            ws = wb.active
            ws.title = 'sheet1'
            ws.append(tableTitle)
            wb.save(filepath)
            time.sleep(3)
        wb = load_workbook(filepath)
        ws = wb.active
        ws.title = 'sheet1'
        for info in allinfo:
            ws.append(info)
        wb.save(filepath)
        print('文件已更新')
    except:
        print('文件更新失败')


cnt = 0
while start_time > end_time:
    url = 'https://m.maoyan.com/mmdb/comments/movie/{}.json?_v_=yes&offset={}&startTime='.format(
        filmid, 15 * cnt) + start_time.replace(' ', '%20')
    cnt += 1
    # print(start_time)
    time.sleep(random.uniform(0.1, 0.5))
    allinfo = getinfo(url, headers)
    if allinfo:
        start_time = allinfo[-1][5]
        start_time = datetime.datetime.strptime(
            start_time, '%Y-%m-%d %H:%M:%S') - datetime.timedelta(seconds=5)
        start_time = datetime.datetime.strftime(start_time,
                                                '%Y-%m-%d %H:%M:%S')
        insertexcel(filepath, allinfo)
    else:
        start_time = datetime.datetime.strptime(
            start_time, '%Y-%m-%d %H:%M:%S') - datetime.timedelta(minutes=5)
        start_time = datetime.datetime.strftime(start_time,
                                                '%Y-%m-%d %H:%M:%S')
